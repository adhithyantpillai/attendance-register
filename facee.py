import cv2
import numpy as np
from openpyxl import load_workbook
from datetime import date
recognizer = cv2.createLBPHFaceRecognizer()
recognizer.load('trainner/trainner.yml')
cascadePath = "haarcascade_frontalface_default.xml"
faceCascade = cv2.CascadeClassifier(cascadePath);
today = date.today()
d= today.day
d =d+5
print d
cam = cv2.VideoCapture(0)
font = cv2.cv.InitFont(cv2.cv.CV_FONT_HERSHEY_SIMPLEX, 1, 1, 0, 1, 1)
while True:
    ret, im =cam.read()
    gray=cv2.cvtColor(im,cv2.COLOR_BGR2GRAY)
    faces=faceCascade.detectMultiScale(gray, 1.2,5)
    for(x,y,w,h) in faces:
        cv2.rectangle(im,(x,y),(x+w,y+h),(225,0,0),2)
        Id, conf = recognizer.predict(gray[y:y+h,x:x+w])
        print Id
        print conf
        if(conf<50):
            if(Id==1):
                Id="adhi"
            elif(Id==2):
                Id="neelima"
                print d
                wb = load_workbook('./students.xlsx')
                sheet = wb.get_sheet_by_name('adhiii')
                mycell= sheet.cell(row=3, column = d)
                
                mycell.value = 'Present'
                wb.save('students.xlsx')
                

            elif(Id==3):
                Id="akash"
        else:
            Id="Unknown"
        cv2.cv.PutText(cv2.cv.fromarray(im),str(Id), (x,y+h),font, 255)
    cv2.imshow('im',im)     
    k = cv2.waitKey(30) & 0xff
    if k == 27: 
        break
cam.release()
cv2.destroyAllWindows()
