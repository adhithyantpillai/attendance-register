from Tkinter import *
import cv2,os
import cv2,os
import numpy as np
from PIL import Image
from tkMessageBox import *

recognizer = cv2.createLBPHFaceRecognizer()
detector= cv2.CascadeClassifier("haarcascade_frontalface_default.xml");

from openpyxl import load_workbook



fields = ('Name', 'Fathers Name', 'Mothers name', 'ROLL no', 'Phn no')
def answer():
    showinfo('Bazinga', 'New student registerd')
def answer2():
    showinfo('cheesee', 'Press Ok for taking picture')
def answer3():
    showinfo('ohoo', 'Relax..Done picturing')
def answer5():
    showinfo('yup', 'Training Completed')



def reg(entries):
 
   wb = load_workbook('./students.xlsx')
   sheet = wb.get_sheet_by_name('adhiii')
   mycell= sheet.cell(row=1, column=8)    
   c = mycell.value
   c = c+1
   mycell.value = c
   
   mycell= sheet.cell(row=c, column=1)
   print c
   data = str(entries['Name'].get())
   mycell.value = data
   
   mycell= sheet.cell(row=c, column=2)
   data = str(entries['Fathers Name'].get())
   mycell.value = data

   mycell= sheet.cell(row=c, column=3)
   data = str(entries['Mothers name'].get())
   mycell.value = data

   mycell= sheet.cell(row=c, column=4)
   data = str(entries['ROLL no'].get())
   mycell.value = data

   mycell= sheet.cell(row=c, column=5)
   data = str(entries['Phn no'].get())
   mycell.value = data
                                            
 
   wb.save('students.xlsx')
   answer()

 

def data(entries):
   import cv2
   cam = cv2.VideoCapture(0)
   detector=cv2.CascadeClassifier('haarcascade_frontalface_default.xml')
   answer2()
   Id = str(entries['ROLL no'].get())
   sampleNum=0
   while(True):
       ret, img = cam.read()
       gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
       faces = detector.detectMultiScale(gray, 1.3, 5)
       for (x,y,w,h) in faces:
           cv2.rectangle(img,(x,y),(x+w,y+h),(255,0,0),2)
        
           #incrementing sample number 
           sampleNum=sampleNum+1
        #saving the captured face in the dataset folder
           cv2.imwrite("dataSet/User."+Id +'.'+ str(sampleNum) + ".jpg", gray[y:y+h,x:x+w])

           cv2.imshow('frame',img)
    #wait for 100 miliseconds 
       if cv2.waitKey(100) & 0xFF == ord('q'):
           break
    # break if the sample number is morethan 20
       elif sampleNum>50:
           break
   cam.release()
   cv2.destroyAllWindows()
   answer3()
def train(void):
  
   faces,Ids = getImagesAndLabels('dataSet')
   recognizer.train(faces, np.array(Ids))
   recognizer.save('trainner/trainner.yml')
   answer5()
   

def getImagesAndLabels(path):
    #get the path of all the files in the folder
    imagePaths=[os.path.join(path,f) for f in os.listdir(path)] 
    #create empth face list
    faceSamples=[]
    #create empty ID list
    Ids=[]
    #now looping through all the image paths and loading the Ids and the images
    for imagePath in imagePaths:

        # Updates in Code
        # ignore if the file does not have jpg extension :
        if(os.path.split(imagePath)[-1].split(".")[-1]!='jpg'):
            continue

        #loading the image and converting it to gray scale
        pilImage=Image.open(imagePath).convert('L')
        #Now we are converting the PIL image into numpy array
        imageNp=np.array(pilImage,'uint8')
        #getting the Id from the image
        Id=int(os.path.split(imagePath)[-1].split(".")[1])
        # extract the face from the training image sample
        faces=detector.detectMultiScale(imageNp)
        #If a face is there then append that in the list as well as Id of it
        for (x,y,w,h) in faces:
            faceSamples.append(imageNp[y:y+h,x:x+w])
            Ids.append(Id)
    return faceSamples,Ids

   


   
def makeform(root, fields):
   entries = {}
   for field in fields:
      row = Frame(root)
      lab = Label(row, width=22, text=field+": ", anchor='w')
      ent = Entry(row)
      ent.insert(0,"0")
      row.pack(side=TOP, fill=X, padx=5, pady=5)
      lab.pack(side=LEFT)
      ent.pack(side=RIGHT, expand=YES, fill=X)
      entries[field] = ent
   return entries



if __name__ == '__main__':
   root = Tk()

   Label(root, text=" ",fg = "black",
		 font = "Times").pack()
   Label(root, text=" ",fg = "black",
		 font = "Times").pack()  
   Label(root, text="Spatez Technolgies",fg = "blue",
		 font = "Times").pack()
   Label(root, text=" ",fg = "black",
		 font = "Times").pack()  
   Label(root, text="Seventh Day School ,Thrissur ",fg = "red",
		 font = "Times").pack()
   Label(root, text="Attendance system",fg = "black",
		 font = "Times").pack()
   Label(root, text=" ",fg = "black",
		 font = "Times").pack()

  
   ents = makeform(root, fields)
   root.bind('<Return>', (lambda event, e=ents: fetch(e)))



   Label(root, text="keep the face parallel to the camera.click train when you are ready",fg = "black").pack()
   Label(root, text="created by adhithyan t pillai ",fg = "black",
		 font = "Times").pack()   
  
   b1 = Button(root, text='   Data   ',
          command=(lambda e=ents: data(e)))
   b1.pack(side=LEFT, padx=100, pady=75)
   b2 = Button(root, text=' Register ',
          command=(lambda e=ents: reg(e)))
   b2.pack(side=LEFT, padx=100, pady=75)
   b3 = Button(root, text=' Training ',
          command=(lambda e=ents: train(e)))
   b3.pack(side=LEFT, padx=100, pady=75)
   root.geometry("800x600")
   
 
   root.mainloop()
